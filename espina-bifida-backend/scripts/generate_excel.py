import sys
import json
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

DARK = "0F172A"
MID  = "1E3A5F"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_AMBER = "FEF3C7"
LIGHT_RED   = "FEE2E2"
HEADER_BG   = "1E3A5F"
HEADER_FG   = "FFFFFF"
ROW_ALT     = "F8FAFC"
ROW_NORM    = "FFFFFF"
ACCENT_1    = "2563EB"
ACCENT_2    = "059669"
ACCENT_3    = "D97706"
ACCENT_4    = "DC2626"

thin = Side(style="thin", color="CBD5E1")
thick = Side(style="medium", color="94A3B8")

def border(all_sides=False, bottom_only=False):
    if bottom_only:
        return Border(bottom=thin)
    s = thin if all_sides else Side(style=None)
    return Border(left=thin, right=thin, top=thin, bottom=thin) if all_sides else Border()

def header_style(ws, row, cols, text, bg=HEADER_BG, fg=HEADER_FG, size=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=fg, size=size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def col_header(ws, row, col, text, bg=MID):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thick)
    ws.row_dimensions[row].height = 18

def data_cell(ws, row, col, value, num_format=None, alt=False, bold=False, color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=9, bold=bold, color=color or "1E293B")
    c.fill = PatternFill("solid", fgColor=ROW_ALT if alt else ROW_NORM)
    c.alignment = Alignment(vertical="center", horizontal="right" if num_format else "left", indent=1)
    c.border = Border(left=thin, right=thin, bottom=thin)
    if num_format:
        c.number_format = num_format
    ws.row_dimensions[row].height = 16
    return c

def kpi_row(ws, row, label, value, color_bg, col_offset=1):
    lc = ws.cell(row=row, column=col_offset, value=label)
    lc.font = Font(name="Arial", size=9, color="475569")
    lc.alignment = Alignment(vertical="center", indent=1)
    lc.fill = PatternFill("solid", fgColor=color_bg)
    lc.border = Border(left=thin, bottom=thin)
    vc = ws.cell(row=row, column=col_offset+1, value=value)
    vc.font = Font(name="Arial", size=10, bold=True, color="0F172A")
    vc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    vc.fill = PatternFill("solid", fgColor=color_bg)
    vc.border = Border(right=thin, bottom=thin)
    ws.row_dimensions[row].height = 17

def add_series_sheet(wb, title, sheet_data, col_headers, col_widths, chart_type, chart_title, y_cols_idx, series_names, chart_anchor, money_cols=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    header_style(ws, 1, len(col_headers), title, size=12)
    for ci, h in enumerate(col_headers, 1):
        col_header(ws, 2, ci, h)

    for ri, row_data in enumerate(sheet_data, 3):
        alt = ri % 2 == 0
        for ci, val in enumerate(row_data, 1):
            fmt = None
            if money_cols and ci in money_cols:
                fmt = '"$"#,##0'
            elif ci > 1:
                fmt = "#,##0"
            data_cell(ws, ri, ci, val, num_format=fmt, alt=alt)

    # totals row
    tr = len(sheet_data) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=tr, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(row=tr, column=1).border = Border(left=thin, bottom=thick)
    for ci in range(2, len(col_headers)+1):
        col_letter = get_column_letter(ci)
        fmt = '"$"#,##0' if money_cols and ci in money_cols else "#,##0"
        c = ws.cell(row=tr, column=ci, value=f"=SUM({col_letter}3:{col_letter}{tr-1})")
        c.font = Font(bold=True, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = Border(right=thin, bottom=thick)
    ws.row_dimensions[tr].height = 18

    # chart
    if chart_type == "bar":
        chart = BarChart()
        chart.grouping = "clustered"
    else:
        chart = LineChart()

    chart.title = chart_title
    chart.style = 10
    chart.height = 10
    chart.width = 18
    chart.y_axis.numFmt = '"$"#,##0' if money_cols else '#,##0'
    chart.y_axis.title = None
    chart.x_axis.title = None

    cats = Reference(ws, min_col=1, min_row=3, max_row=tr-1)
    for idx, sname in zip(y_cols_idx, series_names):
        data_ref = Reference(ws, min_col=idx, min_row=2, max_row=tr-1)
        if chart_type == "bar":
            s = BarChart()
        else:
            s = None
        chart.add_data(data_ref, titles_from_data=True)

    chart.set_categories(cats)
    ws.add_chart(chart, chart_anchor)
    return ws


def generate(stats_json: str) -> bytes:
    data = json.loads(stats_json)
    pacientes     = data.get("pacientes",     {})
    citas         = data.get("citas",         {})
    visitas       = data.get("visitas",       {})
    membresias    = data.get("membresias",    {})
    servicios     = data.get("servicios",     {})
    medicinas     = data.get("medicinas",     {})
    equipo        = data.get("equipo",        {})
    notificaciones= data.get("notificaciones",{})
    series        = data.get("series",        {})

    wb = Workbook()

    # ── 1. RESUMEN ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18

    header_style(ws, 1, 4, "REPORTE MENSUAL — RESUMEN EJECUTIVO", size=13)
    ws.row_dimensions[1].height = 28

    sections = [
        ("PACIENTES", LIGHT_BLUE, [
            ("Total pacientes",        pacientes.get("total", 0)),
            ("Pacientes vivos",        pacientes.get("vivos", 0)),
            ("Pacientes fallecidos",   pacientes.get("fallecidos", 0)),
            ("Nuevos este mes",        pacientes.get("nuevos_mes", 0)),
            ("Con válvula",            pacientes.get("con_valvula", 0)),
            ("Con padecimientos",      pacientes.get("con_padecimientos", 0)),
        ]),
        ("CITAS", LIGHT_GREEN, [
            ("Total citas",            citas.get("total", 0)),
            ("Atendidas",              citas.get("atendidas", 0)),
            ("Canceladas",             citas.get("canceladas", 0)),
            ("Pendientes",             citas.get("pendientes", 0)),
        ]),
        ("VISITAS E INGRESOS", LIGHT_AMBER, [
            ("Total visitas",          visitas.get("total", 0)),
            ("Visitas este mes",       visitas.get("mes", 0)),
            ("Ingresos totales",       visitas.get("ingresos_totales", 0)),
            ("Descuentos totales",     visitas.get("descuentos_totales", 0)),
            ("Ingreso promedio",       visitas.get("ingreso_promedio", 0)),
        ]),
        ("MEMBRESÍAS", LIGHT_BLUE, [
            ("Activas",                membresias.get("activas", 0)),
            ("Inactivas",              membresias.get("inactivas", 0)),
            ("Vencidas",               membresias.get("vencidas", 0)),
        ]),
        ("INVENTARIO — MEDICINAS", LIGHT_GREEN, [
            ("Total medicinas",        medicinas.get("total", 0)),
            ("Stock total",            medicinas.get("stock_total", 0)),
            ("Bajo stock",             medicinas.get("bajo_stock", 0)),
            ("Valor inventario",       medicinas.get("valor_inventario", 0)),
            ("Utilizadas total",       medicinas.get("utilizadas", 0)),
        ]),
        ("INVENTARIO — EQUIPO MÉDICO", LIGHT_AMBER, [
            ("Total equipos",          equipo.get("total", 0)),
            ("En uso",                 equipo.get("en_uso", 0)),
            ("Regresados",             equipo.get("regresados", 0)),
            ("% retorno",              equipo.get("porcentaje_retorno", 0)),
            ("Valor total equipo",     equipo.get("valor_total", 0)),
        ]),
        ("NOTIFICACIONES", LIGHT_RED, [
            ("Este mes",               notificaciones.get("mes", 0)),
            ("Rechazados",             notificaciones.get("rechazados", 0)),
            ("Tasa aprobación (%)",    notificaciones.get("tasa_aprobacion", 0)),
        ]),
    ]

    row_l = row_r = 3
    for i, (name, color, items) in enumerate(sections):
        if i % 2 == 0:  # left (cols 1-2)
            ws.merge_cells(start_row=row_l, start_column=1, end_row=row_l, end_column=2)
            c = ws.cell(row=row_l, column=1, value=name)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=MID)
            c.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            c.border = Border(left=thin, top=thin, bottom=thin)
            ws.cell(row=row_l, column=2).fill = PatternFill("solid", fgColor=MID)
            ws.cell(row=row_l, column=2).border = Border(right=thin, top=thin, bottom=thin)
            ws.row_dimensions[row_l].height = 18
            row_l += 1
            for lbl, val in items:
                kpi_row(ws, row_l, lbl, val, color, col_offset=1)
                row_l += 1
            row_l += 1
        else:           # right (cols 3-4)
            ws.merge_cells(start_row=row_r, start_column=3, end_row=row_r, end_column=4)
            c = ws.cell(row=row_r, column=3, value=name)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=MID)
            c.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            c.border = Border(left=thin, top=thin, bottom=thin)
            ws.cell(row=row_r, column=4).fill = PatternFill("solid", fgColor=MID)
            ws.cell(row=row_r, column=4).border = Border(right=thin, top=thin, bottom=thin)
            ws.row_dimensions[row_r].height = 18
            row_r += 1
            for lbl, val in items:
                kpi_row(ws, row_r, lbl, val, color, col_offset=3)
                row_r += 1
            row_r += 1

    # ── 2. CITAS POR MES ─────────────────────────────────────────────────────
    citas_mes  = series.get("citasMes", [])
    cat_mes    = series.get("citasAtendidasMes", [])
    ccan_mes   = series.get("citasCanceladasMes", [])

    # merge by mes
    def merge_by_mes(*lists_keys):
        merged = {}
        for lst, key in lists_keys:
            for r in lst:
                m = r.get("mes", "")
                if m not in merged:
                    merged[m] = {"mes": m}
                merged[m][key] = r.get("total", 0)
        return [merged[k] for k in sorted(merged.keys())]

    citas_data = merge_by_mes((citas_mes, "total"), (cat_mes, "atendidas"), (ccan_mes, "canceladas"))
    rows_citas = [[r["mes"], r.get("total",0), r.get("atendidas",0), r.get("canceladas",0)] for r in citas_data]

    ws_c = wb.create_sheet("Citas por Mes")
    ws_c.sheet_view.showGridLines = False
    for ci, w in enumerate([18,14,14,14], 1):
        ws_c.column_dimensions[get_column_letter(ci)].width = w
    header_style(ws_c, 1, 4, "CITAS POR MES", size=12)
    for ci, h in enumerate(["Mes","Total","Atendidas","Canceladas"], 1):
        col_header(ws_c, 2, ci, h)
    for ri, rd in enumerate(rows_citas, 3):
        alt = ri % 2 == 0
        data_cell(ws_c, ri, 1, rd[0], alt=alt)
        for ci in range(2, 5):
            data_cell(ws_c, ri, ci, rd[ci-1], num_format="#,##0", alt=alt)
    tr = len(rows_citas) + 3
    ws_c.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws_c.cell(row=tr, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws_c.cell(row=tr, column=1).border = Border(left=thin, bottom=thick)
    for ci in range(2, 5):
        cl = get_column_letter(ci)
        c = ws_c.cell(row=tr, column=ci, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.font = Font(bold=True, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = Border(right=thin, bottom=thick)

    chart_c = BarChart()
    chart_c.type = "col"
    chart_c.grouping = "clustered"
    chart_c.title = "Citas por mes"
    chart_c.style = 10
    chart_c.height = 11
    chart_c.width = 20
    chart_c.y_axis.numFmt = "#,##0"
    cats_c = Reference(ws_c, min_col=1, min_row=3, max_row=tr-1)
    data_c = Reference(ws_c, min_col=2, max_col=4, min_row=2, max_row=tr-1)
    chart_c.add_data(data_c, titles_from_data=True)
    chart_c.set_categories(cats_c)
    ws_c.add_chart(chart_c, f"A{tr+3}")

    # ── 3. INGRESOS POR MES ───────────────────────────────────────────────────
    ing_mes  = series.get("ingresosMes", [])
    desc_mes = series.get("descuentosMes", [])
    ing_data = merge_by_mes((ing_mes, "ingresos"), (desc_mes, "descuentos"))
    rows_ing = [[r["mes"], r.get("ingresos",0), r.get("descuentos",0)] for r in ing_data]

    ws_i = wb.create_sheet("Ingresos por Mes")
    ws_i.sheet_view.showGridLines = False
    for ci, w in enumerate([18,18,18,18], 1):
        ws_i.column_dimensions[get_column_letter(ci)].width = w
    header_style(ws_i, 1, 4, "INGRESOS Y DESCUENTOS POR MES", size=12)
    for ci, h in enumerate(["Mes","Ingresos","Descuentos","Neto"], 1):
        col_header(ws_i, 2, ci, h)
    for ri, rd in enumerate(rows_ing, 3):
        alt = ri % 2 == 0
        data_cell(ws_i, ri, 1, rd[0], alt=alt)
        data_cell(ws_i, ri, 2, rd[1], num_format='"$"#,##0', alt=alt)
        data_cell(ws_i, ri, 3, rd[2], num_format='"$"#,##0', alt=alt)
        col_b, col_c = get_column_letter(2), get_column_letter(3)
        c = ws_i.cell(row=ri, column=4, value=f"=B{ri}-C{ri}")
        c.number_format = '"$"#,##0'
        c.font = Font(name="Arial", size=9, bold=True)
        c.fill = PatternFill("solid", fgColor=ROW_ALT if alt else ROW_NORM)
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = Border(right=thin, bottom=thin)
    tr_i = len(rows_ing) + 3
    ws_i.cell(row=tr_i, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws_i.cell(row=tr_i, column=1).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws_i.cell(row=tr_i, column=1).border = Border(left=thin, bottom=thick)
    for ci, col in enumerate(range(2, 5), 2):
        cl = get_column_letter(ci)
        val = f"=SUM({cl}3:{cl}{tr_i-1})" if ci < 4 else f"=B{tr_i}-C{tr_i}"
        c = ws_i.cell(row=tr_i, column=ci, value=val)
        c.font = Font(bold=True, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = Border(right=thin, bottom=thick)

    chart_i = BarChart()
    chart_i.type = "col"
    chart_i.grouping = "clustered"
    chart_i.title = "Ingresos vs Descuentos"
    chart_i.style = 10
    chart_i.height = 11
    chart_i.width = 20
    chart_i.y_axis.numFmt = '"$"#,##0'
    cats_i = Reference(ws_i, min_col=1, min_row=3, max_row=tr_i-1)
    data_i = Reference(ws_i, min_col=2, max_col=3, min_row=2, max_row=tr_i-1)
    chart_i.add_data(data_i, titles_from_data=True)
    chart_i.set_categories(cats_i)
    ws_i.add_chart(chart_i, f"A{tr_i+3}")

    # ── 4. SERVICIOS POR MES ──────────────────────────────────────────────────
    vis_mes  = series.get("visitasMes", [])
    srv_mes  = series.get("serviciosMes", [])
    med_mes  = series.get("medicinasUtilizadasMes", [])
    srv_data = merge_by_mes((vis_mes, "visitas"), (srv_mes, "servicios"), (med_mes, "medicinas"))
    rows_srv = [[r["mes"], r.get("visitas",0), r.get("servicios",0), r.get("medicinas",0)] for r in srv_data]

    ws_s = wb.create_sheet("Servicios por Mes")
    ws_s.sheet_view.showGridLines = False
    for ci, w in enumerate([18,14,14,14], 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    header_style(ws_s, 1, 4, "SERVICIOS, VISITAS Y MEDICINAS POR MES", size=12)
    for ci, h in enumerate(["Mes","Visitas","Servicios","Medicinas"], 1):
        col_header(ws_s, 2, ci, h)
    for ri, rd in enumerate(rows_srv, 3):
        alt = ri % 2 == 0
        data_cell(ws_s, ri, 1, rd[0], alt=alt)
        for ci in range(2, 5):
            data_cell(ws_s, ri, ci, rd[ci-1], num_format="#,##0", alt=alt)
    tr_s = len(rows_srv) + 3
    ws_s.cell(row=tr_s, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws_s.cell(row=tr_s, column=1).fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    ws_s.cell(row=tr_s, column=1).border = Border(left=thin, bottom=thick)
    for ci in range(2, 5):
        cl = get_column_letter(ci)
        c = ws_s.cell(row=tr_s, column=ci, value=f"=SUM({cl}3:{cl}{tr_s-1})")
        c.font = Font(bold=True, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=LIGHT_AMBER)
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = Border(right=thin, bottom=thick)

    chart_s = LineChart()
    chart_s.title = "Servicios, Visitas y Medicinas"
    chart_s.style = 10
    chart_s.height = 11
    chart_s.width = 20
    cats_s = Reference(ws_s, min_col=1, min_row=3, max_row=tr_s-1)
    data_s = Reference(ws_s, min_col=2, max_col=4, min_row=2, max_row=tr_s-1)
    chart_s.add_data(data_s, titles_from_data=True)
    chart_s.set_categories(cats_s)
    ws_s.add_chart(chart_s, f"A{tr_s+3}")

    # ── 5. PACIENTES POR MES ──────────────────────────────────────────────────
    pac_mes = series.get("pacientesNuevosMes", [])
    rows_pac = [[r.get("mes",""), r.get("total",0)] for r in pac_mes]

    ws_p = wb.create_sheet("Pacientes Nuevos")
    ws_p.sheet_view.showGridLines = False
    ws_p.column_dimensions["A"].width = 18
    ws_p.column_dimensions["B"].width = 18
    header_style(ws_p, 1, 2, "PACIENTES NUEVOS POR MES", size=12)
    col_header(ws_p, 2, 1, "Mes")
    col_header(ws_p, 2, 2, "Pacientes Nuevos")
    for ri, rd in enumerate(rows_pac, 3):
        alt = ri % 2 == 0
        data_cell(ws_p, ri, 1, rd[0], alt=alt)
        data_cell(ws_p, ri, 2, rd[1], num_format="#,##0", alt=alt)
    tr_p = len(rows_pac) + 3
    ws_p.cell(row=tr_p, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws_p.cell(row=tr_p, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws_p.cell(row=tr_p, column=1).border = Border(left=thin, bottom=thick)
    c = ws_p.cell(row=tr_p, column=2, value=f"=SUM(B3:B{tr_p-1})")
    c.font = Font(bold=True, name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.number_format = "#,##0"
    c.alignment = Alignment(horizontal="right", indent=1)
    c.border = Border(right=thin, bottom=thick)

    chart_p = BarChart()
    chart_p.type = "col"
    chart_p.title = "Pacientes Nuevos por Mes"
    chart_p.style = 10
    chart_p.height = 11
    chart_p.width = 16
    cats_p = Reference(ws_p, min_col=1, min_row=3, max_row=tr_p-1)
    data_p = Reference(ws_p, min_col=2, min_row=2, max_row=tr_p-1)
    chart_p.add_data(data_p, titles_from_data=True)
    chart_p.set_categories(cats_p)
    ws_p.add_chart(chart_p, f"A{tr_p+3}")

    # ── 6. INVENTARIO ─────────────────────────────────────────────────────────
    ws_inv = wb.create_sheet("Inventario")
    ws_inv.sheet_view.showGridLines = False
    ws_inv.column_dimensions["A"].width = 32
    ws_inv.column_dimensions["B"].width = 20
    header_style(ws_inv, 1, 2, "INVENTARIO — MEDICINAS Y EQUIPO", size=12)

    inv_rows = [
        ("— MEDICINAS —", None, MID),
        ("Total medicinas",       medicinas.get("total",0),              LIGHT_BLUE),
        ("Stock total",           medicinas.get("stock_total",0),         LIGHT_BLUE),
        ("Bajo stock (<10)",      medicinas.get("bajo_stock",0),          LIGHT_AMBER),
        ("Valor inventario",      medicinas.get("valor_inventario",0),    LIGHT_GREEN),
        ("Medicinas utilizadas",  medicinas.get("utilizadas",0),          LIGHT_BLUE),
        ("Actualizaciones inv.",  medicinas.get("actualizaciones_inventario",0), LIGHT_BLUE),
        ("— EQUIPO MÉDICO —", None, MID),
        ("Total equipos",         equipo.get("total",0),                  LIGHT_BLUE),
        ("Cantidad disponible",   equipo.get("cantidad_total",0),         LIGHT_BLUE),
        ("En uso",                equipo.get("en_uso",0),                 LIGHT_AMBER),
        ("Regresados",            equipo.get("regresados",0),             LIGHT_GREEN),
        ("% retorno",             equipo.get("porcentaje_retorno",0),     LIGHT_BLUE),
        ("Valor total equipo",    equipo.get("valor_total",0),            LIGHT_GREEN),
    ]
    money_set = {"Valor inventario", "Valor total equipo"}
    for ri, item in enumerate(inv_rows, 3):
        lbl, val, color = item
        if val is None:
            ws_inv.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=2)
            c = ws_inv.cell(row=ri, column=1, value=lbl)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(indent=1, vertical="center")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws_inv.row_dimensions[ri].height = 18
        else:
            fmt = '"$"#,##0' if lbl in money_set else "#,##0"
            kpi_row(ws_inv, ri, lbl, val, color, col_offset=1)
            ws_inv.cell(row=ri, column=2).number_format = fmt

    # ── 7. NOTIFICACIONES ─────────────────────────────────────────────────────
    notif_mes = series.get("notificacionesMes", [])
    rows_notif = [[r.get("mes",""), r.get("total",0)] for r in notif_mes]

    ws_n = wb.create_sheet("Notificaciones")
    ws_n.sheet_view.showGridLines = False
    ws_n.column_dimensions["A"].width = 18
    ws_n.column_dimensions["B"].width = 20
    header_style(ws_n, 1, 2, "NOTIFICACIONES POR MES", size=12)
    col_header(ws_n, 2, 1, "Mes")
    col_header(ws_n, 2, 2, "Total Notificaciones")
    for ri, rd in enumerate(rows_notif, 3):
        alt = ri % 2 == 0
        data_cell(ws_n, ri, 1, rd[0], alt=alt)
        data_cell(ws_n, ri, 2, rd[1], num_format="#,##0", alt=alt)
    tr_n = len(rows_notif) + 3
    ws_n.cell(row=tr_n, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws_n.cell(row=tr_n, column=1).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws_n.cell(row=tr_n, column=1).border = Border(left=thin, bottom=thick)
    c = ws_n.cell(row=tr_n, column=2, value=f"=SUM(B3:B{tr_n-1})")
    c.font = Font(bold=True, name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=LIGHT_RED)
    c.number_format = "#,##0"
    c.alignment = Alignment(horizontal="right", indent=1)
    c.border = Border(right=thin, bottom=thick)

    chart_n = BarChart()
    chart_n.type = "col"
    chart_n.title = "Notificaciones por Mes"
    chart_n.style = 10
    chart_n.height = 10
    chart_n.width = 14
    cats_n = Reference(ws_n, min_col=1, min_row=3, max_row=tr_n-1)
    data_n = Reference(ws_n, min_col=2, min_row=2, max_row=tr_n-1)
    chart_n.add_data(data_n, titles_from_data=True)
    chart_n.set_categories(cats_n)
    ws_n.add_chart(chart_n, f"A{tr_n+3}")

    # ── save ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    raw = sys.stdin.read()
    result = generate(raw)
    sys.stdout.buffer.write(base64.b64encode(result))