import base64
import json
import subprocess
import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT_DIR))

from scripts import generate_excel


def sample_stats():
    return {
        "pacientes": {
            "total": 10,
            "vivos": 9,
            "fallecidos": 1,
            "nuevos_mes": 2,
            "con_valvula": 3,
            "con_padecimientos": 8,
        },
        "citas": {
            "total": 20,
            "atendidas": 15,
            "canceladas": 3,
            "pendientes": 2,
        },
        "visitas": {
            "total": 30,
            "mes": 5,
            "ingresos_totales": 10000,
            "descuentos_totales": 1000,
            "ingreso_promedio": 300,
        },
        "membresias": {
            "activas": 7,
            "inactivas": 2,
            "vencidas": 1,
        },
        "medicinas": {
            "total": 12,
            "stock_total": 100,
            "bajo_stock": 2,
            "valor_inventario": 5000,
            "utilizadas": 25,
            "actualizaciones_inventario": 4,
        },
        "equipo": {
            "total": 6,
            "cantidad_total": 14,
            "en_uso": 3,
            "regresados": 2,
            "porcentaje_retorno": 66,
            "valor_total": 12000,
        },
        "notificaciones": {
            "mes": 4,
            "rechazados": 1,
            "tasa_aprobacion": 75,
        },
        "series": {
            "citasMes": [
                {"mes": "2026-05", "total": 10},
                {"mes": "2026-06", "total": 20},
            ],
            "citasAtendidasMes": [
                {"mes": "2026-05", "total": 8},
                {"mes": "2026-06", "total": 15},
            ],
            "citasCanceladasMes": [
                {"mes": "2026-05", "total": 1},
                {"mes": "2026-06", "total": 3},
            ],
            "ingresosMes": [
                {"mes": "2026-05", "total": 5000},
                {"mes": "2026-06", "total": 7000},
            ],
            "descuentosMes": [
                {"mes": "2026-05", "total": 500},
                {"mes": "2026-06", "total": 800},
            ],
            "visitasMes": [
                {"mes": "2026-05", "total": 4},
                {"mes": "2026-06", "total": 6},
            ],
            "serviciosMes": [
                {"mes": "2026-05", "total": 3},
                {"mes": "2026-06", "total": 5},
            ],
            "medicinasUtilizadasMes": [
                {"mes": "2026-05", "total": 10},
                {"mes": "2026-06", "total": 15},
            ],
            "pacientesNuevosMes": [
                {"mes": "2026-05", "total": 1},
                {"mes": "2026-06", "total": 2},
            ],
            "notificacionesMes": [
                {"mes": "2026-05", "total": 2},
                {"mes": "2026-06", "total": 4},
            ],
        },
    }


def test_generate_excel_returns_xlsx_bytes():
    result = generate_excel.generate(json.dumps(sample_stats()))

    assert isinstance(result, bytes)
    assert result[:2] == b"PK"  # los xlsx son archivos zip internamente


def test_generate_excel_creates_expected_sheets():
    result = generate_excel.generate(json.dumps(sample_stats()))
    wb = load_workbook(BytesIO(result), data_only=False)

    expected_sheets = [
        "Resumen",
        "Citas por Mes",
        "Ingresos por Mes",
        "Servicios por Mes",
        "Pacientes Nuevos",
        "Inventario",
        "Notificaciones",
    ]

    assert wb.sheetnames == expected_sheets


def test_generate_excel_summary_contains_main_title_and_kpis():
    result = generate_excel.generate(json.dumps(sample_stats()))
    wb = load_workbook(BytesIO(result), data_only=False)
    ws = wb["Resumen"]

    assert ws["A1"].value == "REPORTE MENSUAL — RESUMEN EJECUTIVO"
    assert ws["A3"].value == "PACIENTES"
    assert ws["A4"].value == "Total pacientes"
    assert ws["B4"].value == 10


def test_generate_excel_citas_sheet_contains_data_and_total_formula():
    result = generate_excel.generate(json.dumps(sample_stats()))
    wb = load_workbook(BytesIO(result), data_only=False)
    ws = wb["Citas por Mes"]

    assert ws["A1"].value == "CITAS POR MES"
    assert ws["A2"].value == "Mes"
    assert ws["B2"].value == "Total"

    assert ws["A3"].value == "2026-05"
    assert ws["B3"].value == 10
    assert ws["C3"].value == 8
    assert ws["D3"].value == 1

    assert ws["A5"].value == "TOTAL"
    assert ws["B5"].value == "=SUM(B3:B4)"
    assert ws["C5"].value == "=SUM(C3:C4)"
    assert ws["D5"].value == "=SUM(D3:D4)"


def test_generate_excel_ingresos_sheet_calculates_neto_formula():
    result = generate_excel.generate(json.dumps(sample_stats()))
    wb = load_workbook(BytesIO(result), data_only=False)
    ws = wb["Ingresos por Mes"]

    assert ws["A1"].value == "INGRESOS Y DESCUENTOS POR MES"
    assert ws["B3"].value == 5000
    assert ws["C3"].value == 500
    assert ws["D3"].value == "=B3-C3"


def test_generate_excel_works_with_empty_payload():
    result = generate_excel.generate(json.dumps({}))
    wb = load_workbook(BytesIO(result), data_only=False)

    assert "Resumen" in wb.sheetnames
    assert "Inventario" in wb.sheetnames
    assert wb["Resumen"]["A1"].value == "REPORTE MENSUAL — RESUMEN EJECUTIVO"


def test_generate_excel_raises_error_with_invalid_json():
    with pytest.raises(json.JSONDecodeError):
        generate_excel.generate("{json-invalido}")


def test_generate_excel_cli_outputs_base64_xlsx():
    payload = json.dumps(sample_stats())
    script_path = Path(generate_excel.__file__).resolve()

    completed = subprocess.run(
        [sys.executable, str(script_path)],
        input=payload,
        text=True,
        capture_output=True,
        check=True,
    )

    decoded = base64.b64decode(completed.stdout)

    assert decoded[:2] == b"PK"